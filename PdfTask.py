import pdfkit
import sys
import csv
import os
import openpyxl
import doctest
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl.styles import Font, NamedStyle, Side, Border
from jinja2 import Environment, FileSystemLoader
from collections import OrderedDict


class Vacancy:
    """
    Класс представления вакансии.

    Attributes:
        name (str): Название вакансии
        salary (Salary): Представление запрлаты
        areaName (str): Название города
        publishedAt (str): Дата публикации
    """

    def __init__(self, name, salary, areaName, publishedAt):
        """
        Инициализирует объект Vacancy

        Args:
            name (str): Название вакансии
            salary (Salary): Представление запрлаты
            areaName (str): Название города
            publishedAt (str): Дата публикации
        """
        self.name, self.salary, self.areaName, self.publishedAt = name, salary, areaName, publishedAt


class DataSet:
    """
    Класс, отвечающий за чтение и подготовку данных из CSV-файла.

    Attributes:
        fileName (str): Название файла
        correctFields (list[str]): Поля необходимые для инициализации вакансии
        vacancyNameParameter (str): Название выбранной профессии
        vacanciesByYear (dict): Общая динамика вакансий по годам
        vacancyByYear (dict): Динамика вакансий по годам для выбранной профессии
        vacanciesByArea (dict): Общая динамика вакансий по городам
        vacanciesObjects (list[Vacancy]): Список всех вакансий
        period (list[int]): Период лет
    """
    correctFields = ["name", "salary_from", "area_name", "published_at"]

    def __init__(self, fileName, vacancyNameParameter):
        """
        Инициализирует объект DataSet

        Args:
            fileName (str): Название файла
            vacancyNameParameter (str): Название выбранной профессии
        """
        self.fileName = fileName
        self.vacancyNameParameter = vacancyNameParameter
        self.vacanciesObjects = self.__UniversalParserCSV(fileName)
        self.vacanciesByYear = self.__VacancyFilterByYear()
        self.vacancyByYear = self.__VacancyFilterByYear(self.vacancyNameParameter)
        self.vacanciesByArea = self.__VacancyFilterByArea()

    def __UniversalParserCSV(self, fileName):
        """
        Парсит CSV файл по вакансиям в список

        Args:
            fileName (str): Название файла

        Returns:
            list[Vacancy]: Список всех вакансий
        """
        fileReader, columnNames = self.__CsvReader(fileName)
        vacanciesObjects = self.__CsvFilter(fileReader, columnNames)
        return vacanciesObjects

    def __CsvReader(self, fileName):
        """
        Считывает CSV файл.
        Если файл пустой - выводит строку "Пустой файл" и прерывает работу программы,
        иначе - возвращает все строки из файла в виде OrderedDict и список заголовков полей

        Args:
            fileName (str): Название файла
        """
        file = open(fileName, encoding='utf-8-sig', newline='')
        if os.stat(fileName).st_size == 0:
            print("Пустой файл")
            sys.exit()
        fileReader = csv.DictReader(file)
        columnNames = fileReader.fieldnames
        return fileReader, columnNames

    def __CsvFilter(self, fileReader, columnNames):
        """
        Обрабатывает полученные на вход словари, возвращает список всех вакансий

        Args:
            fileReader: Все строки из файла в виде словарей
            columnNames: Список заголовков полей

        Returns:
            list[Vacancy]: Список всех вакансий
        """
        vacancies = []
        columnsCount = len(columnNames)
        for row in fileReader:
            if all(row.values()) and columnsCount == len(row):
                tempRow = {name: row[name] for name in columnNames}
                tempRow['salary_from'] = Salary(tempRow['salary_from'], tempRow.pop('salary_to'),
                                                tempRow.pop("salary_currency"))
                vacancies.append(Vacancy(*[tempRow[key] for key in self.correctFields]))
        return vacancies

    def DynamicsSalaries(self):
        """
        Обрабатывает общую динамику вакансий по годам,
        возвращает динамику уровня зарплат по годам

        Returns:
            dict: Динамика уровня зарплат по годам
        """
        dinamicsSalaries = self.vacanciesByYear.copy()
        for dataByYear in dinamicsSalaries.items():
            averagesByYear = [vacancy.salary.GetAverage() for vacancy in dataByYear[1]]
            dinamicsSalaries[dataByYear[0]] = int(sum(averagesByYear) / len(averagesByYear))
        return dinamicsSalaries

    def DynamicsCountVacancies(self):
        """
        Обрабатывает общую динамику вакансий по годам,
        возвращает динамику количества вакансий по годам

        Returns:
            dict: Динамика количества вакансий по годам
        """
        vacancyCountByYear = self.vacanciesByYear.copy()
        for dataByYear in vacancyCountByYear.items():
            vacancyCountByYear[dataByYear[0]] = len(dataByYear[1])
        return vacancyCountByYear

    def DynamicsSalariesAtVacancy(self):
        """
        Обрабатывает динамику вакансий по годам для выбранной профессии,
        возвращает динамику уровня зарплат по годам для выбранной профессии

        Returns:
            dict: Динамика уровня зарплат по годам для выбранной профессии
        """
        dinamicsSalaries = self.vacancyByYear.copy()
        if not dinamicsSalaries:
            return {key: 0 for key in self.period}
        for dataByYear in dinamicsSalaries.items():
            averagesByYear = [vacancy.salary.GetAverage() for vacancy in dataByYear[1]]
            dinamicsSalaries[dataByYear[0]] = int(sum(averagesByYear) / len(averagesByYear))
        return dinamicsSalaries

    def DynamicsCountVacanciesAtVacancy(self):
        """
        Обрабатывает динамику вакансий по годам для выбранной профессии,
        возвращает динамику количества вакансий по годам для выбранной профессии

        Returns:
            dict: Динамика уровня зарплат по годам для выбранной профессии
        """
        vacancyCountByYear = self.vacancyByYear.copy()
        if not vacancyCountByYear:
            return {key: 0 for key in self.period}
        for dataByYear in vacancyCountByYear.items():
            vacancyCountByYear[dataByYear[0]] = len(dataByYear[1])
        return vacancyCountByYear

    def CitiesSalaryLevel(self):
        """
        Обрабатывает общую динамику вакансий по городам,
        возвращает динамику уровня зарплат по городам (в порядке убывания, первые 10 значений)

        Returns:
            dict: Динамика уровня зарплат по городам
        """
        vacanciesByArea = self.vacanciesByArea.copy()
        for dataByArea in vacanciesByArea.items():
            averagesByArea = [vacancy.salary.GetAverage() for vacancy in dataByArea[1]]
            vacanciesByArea[dataByArea[0]] = int(sum(averagesByArea) / len(averagesByArea))
        vacanciesByArea = list(
            {k: v for k, v in sorted(vacanciesByArea.items(), key=lambda item: item[1], reverse=True)}.items())
        vacanciesByArea = {items[0]: items[1] for items in vacanciesByArea}
        return vacanciesByArea

    def CitiesRatioVacancies(self):
        """
        Обрабатывает общую динамику вакансий по городам,
        возвращает динамику доли вакансий по городам (в порядке убывания, первые 10 значений)

        Returns:
            dict: Доля вакансий по городам (в порядке убывания)
        """
        vacanciesByArea = self.vacanciesByArea.copy()
        for dataByArea in vacanciesByArea.items():
            vacanciesByArea[dataByArea[0]] = round(len(dataByArea[1]) / len(self.vacanciesObjects), 4)
        vacanciesByArea = list(
            {k: v for k, v in sorted(vacanciesByArea.items(), key=lambda item: item[1], reverse=True)}.items())
        vacanciesByArea = {items[0]: items[1] for items in vacanciesByArea}
        return vacanciesByArea

    def __VacancyFilterByYear(self, vacancyName=None):
        """
        Обрабатывает список всех вакансий,
        если указана профессиия - возвращает динамику вакансий по годам для выбранной профессии,
        иначе - возвращает общую динамику профессий по годам

        Args:
            vacancyName: Название выбранной профессии

        Returns:
            dict: Динамика вакансий по годам
        """
        dinamicsSalaries = {}
        self.period = []
        for vacancy in self.vacanciesObjects:
            year = datetime.strptime(vacancy.publishedAt, '%Y-%m-%dT%H:%M:%S%z').year
            self.period.append(year)
            if vacancyName is None or vacancyName in vacancy.name:
                dinamicsSalaries.setdefault(year, []).append(vacancy)
        return dinamicsSalaries

    def __VacancyFilterByArea(self):
        """
        Обрабатывает список всех вакансий,
        возвращает общую динамику вакансий по городам

        Returns:
            dict: Общая динамика вакансий по годам
        """
        vacanciesByArea = {}
        for vacancy in self.vacanciesObjects:
            vacanciesByArea.setdefault(vacancy.areaName, []).append(vacancy)
        vacanciesByArea = self.__ClearByArea(vacanciesByArea)

        return vacanciesByArea

    def __ClearByArea(self, vacanciesByArea):
        """
        Обрабатывает общую динамику вакансий по городам,
        возваращает общую динамику вакансий только тех городов,
        в которых кол-во вакансий больше или равно 1% от общего числа вакансий
        (при вычислении 1% применяется округление вниз)

        Args:
            vacanciesByArea: Общая динамика вакансий по городам

        Returns:
            dict: Общая динамика вакансий по городам
        """
        tempAreas = vacanciesByArea.copy()
        for keyArea in vacanciesByArea.keys():
            if len(vacanciesByArea[keyArea]) / len(self.vacanciesObjects) < 0.01:
                tempAreas.pop(keyArea)
        return tempAreas


class Salary:
    """
    Класс представления зарплаты.

    Attributes:
        salaryFrom (int): Зарплато от
        salaryTo (int): Зарплата до
        salaryCurrency (str): Название валюты
        currencyToRub (dict): Курс обмена валют
    """

    currencyToRub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, salaryFrom, salaryTo, salaryCurrency):
        """
        Инициализирует объект Salary

        Args:
            salaryFrom (int): Зарплато от
            salaryTo (int): Зарплата до
            salaryCurrency (str): Название валюты
        """
        self.salaryFrom, self.salaryTo, self.salaryCurrency = salaryFrom, salaryTo, salaryCurrency

    def ChangeCurrency(self, salary):
        """
        Конвертирует сумму из любой валюты в рубли

        Args:
            salary(float): Сумма

        Returns:
            int: Сумма в рублях
        """
        return salary * self.currencyToRub[self.salaryCurrency]

    def GetAverage(self):
        """
        Считает среднюю зарплату в рублях по формуле (salaryFrom + salaryTo) / 2

        Returns:
            int: Средняя зарплата в рублях

        >>> Salary(100, 200, "RUR").GetAverage()
        150.0
        >>> Salary(0, 200, "RUR").GetAverage()
        100.0
        >>> Salary(200, 0, "RUR").GetAverage()
        100.0
        >>> Salary(0, 0, "RUR").GetAverage()
        0.0
        >>> Salary(100, 100, "RUR").GetAverage()
        100.0
        """
        return self.ChangeCurrency((int(float(self.salaryFrom)) + int(float(self.salaryTo))) / 2)


class InputConnect:
    """
    Класс, отвечающий за обработку параметров вводимых пользователем: название файла, название профессии,
    а также вывож данных для таблицы на экран

    Attributes:
        __requests (dict): Словарь запросов данных пользователю
        _responses (dict): Словарь выходных данных пользователю
        fileName (str): Название файла
        vacancyName (str): Название выбранной профессии
    """

    __requests = {"Введите название файла: ": lambda fileName: fileName,
                  "Введите название профессии: ": lambda vacancyName: vacancyName}

    _responses = {"Динамика уровня зарплат по годам: ": lambda dataSet: dataSet.DynamicsSalaries(),
                  "Динамика уровня зарплат по годам для выбранной профессии: ": lambda
                      dataSet: dataSet.DynamicsSalariesAtVacancy(),
                  "Динамика количества вакансий по годам: ": lambda dataSet: dataSet.DynamicsCountVacancies(),
                  "Динамика количества вакансий по годам для выбранной профессии: ": lambda
                      dataSet: dataSet.DynamicsCountVacanciesAtVacancy(),
                  "Уровень зарплат по городам (в порядке убывания): ": lambda dataSet: dataSet.CitiesSalaryLevel(),
                  "Доля вакансий по городам (в порядке убывания): ": lambda dataSet: dataSet.CitiesRatioVacancies()}

    def __init__(self):
        """
        Инициализирует объект InputConnect
        """
        self.fileName, self.vacancyName = self.__GetData()

    def __GetData(self):
        """
        Запрашивает у пользователя все необходимые данные

        Returns:
            dict_values(list[str]): Данные введенные пользователем
        """
        data = {}
        for request in self.__requests.keys():
            data[request] = input(request)
        return data.values()

    def PrintData(self, dataSet):
        """
        Выводит данные статистики вакансий в виде "Название динамики: соответствующая динамика"

        Args:
            dataSet (DataSet): Данные файла
        """

        dataSet = self.GetListData(dataSet)
        for i, response in enumerate(self._responses.keys()):
            if i < 4:
                print(f'{response}{dataSet[i]}')
            else:
                outputData = {k: v for k, v in list(dataSet[i].items())[:10]}
                print(f'{response}{outputData}')

    def GetListData(self, dataSet):
        """
        Преобразует данные из DataSet в данные соответсвующих динамик

        Args:
            dataSet (DataSet): Данные файла

        Returns:
            list[dict]: Данные соответсвующих динамик
        """
        data = []
        for response in self._responses.items():
            data.append(response[1](dataSet))
        return data


class Report:
    """
    Класс, формирующий отчет для пользователя
    """

    def __init__(self, vacancyName):
        """
        Инициализирует объект Report

        Args:
            vacancyName (str): название профессий
        """

        self.vacancyName = vacancyName

    def CheckEmptyText(self, value):
        """
        Проверяет ячейку на пустоту,
        если ячейка пустая (None) - задает ячейке значение ""

        Args:
            value (str): Ячейка таблицы

        Returns:
            str: значение ячейки
        """
        if value is None:
            return ""
        return str(value)

    def __CompleteSheet(self, sheet, data, indexesData, count=None):
        """
        Заполняет таблицу данными из списка, в указанном диапазоне.
        Если указано количество необходимых данных, отсчитывает количество от начала списка

        Args:
            sheet (Worksheet): Excel страница (таблица)
            data list[dict]: Данные
            indexesData (tuple): Диапазон данных
            count (int): Количество данных

        Returns:
            Таблица с данными из списка, в указанном диапазоне.
        """
        i, j = indexesData
        if count is None:
            count = len(data[i].keys()) + 1
        for year in list(data[i].keys())[:count]:
            row = [year]
            for dictData in data[i:j]:
                row.append(dictData[year])
            sheet.append(row)
        return sheet

    def __CopySheetToSheet(self, sheet1, sheet2):
        """
        Копирует содержимое одной таблицы  в другую с отступом в один столбец

        Args:
            sheet1 (Worksheet): Копируемая таблица
            sheet2 (Worksheet): Таблица, в которую необходимо скопировать данные

        Returns:
            Таблица с данными из другой таблицы
        """
        maxColumn = sheet2.max_column
        for i in range(1, sheet1.max_row + 1):
            for j in range(1, sheet1.max_column + 1):
                sheet2.cell(row=i + 1, column=maxColumn + j - 2).value = sheet1.cell(row=i, column=j).value
        return sheet2

    def __StyleSheet(self, sheet, headingsStyle, cellStyle):
        """
        Устанавливает стили заголовков и ячеек для таблицы

        Args:
            sheet (Worksheet): Таблица
            headingsStyle (NamedStyle): Стиль заголовков
            cellStyle (NamedStyle): стиль ячеек
        """

        for row in sheet.rows:
            for cell in row:
                cell.style = cellStyle
        for cell in sheet["1:1"]:
            cell.style = headingsStyle
        for column in sheet.columns:
            length = max(len(self.CheckEmptyText(cell.value)) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = length + 2

    def GenerateExcel(self, listData):
        """
        Формирует Excel файл "report.xlsx", в котором представлены динамики данных из файла в виде таблиц

        Args:
            listData list[dict]: Данные файла
        """
        book = openpyxl.Workbook()
        book.remove(book.active)
        sheet1 = book.create_sheet("Статистика по годам")
        sheet2 = book.create_sheet("Статистика по городам")
        headingsByYear = ["Год", "Средняя зарплата", f'Средняя зарплата - {self.vacancyName}', "Количество вакансий",
                          f'Количество вакансий - {self.vacancyName}']
        headingsByCity = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        sheet1.append(headingsByYear)
        sheet2.append(headingsByCity)

        headingsStyle = NamedStyle(name='headingsStyle')
        headingsStyle.font = Font(bold=True)
        headingsStyle.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                      bottom=Side(style='thin'))
        cellStyle = NamedStyle(name='cellStyle')
        cellStyle.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                  bottom=Side(style='thin'))

        tempSheet = book.create_sheet("tempSheet")
        sheet1 = self.__CompleteSheet(sheet1, listData, (0, 4))
        sheet2 = self.__CompleteSheet(sheet2, listData, (4, 5), 10)
        tempSheet = self.__CompleteSheet(tempSheet, listData, (5, 6), 10)
        sheet2 = self.__CopySheetToSheet(tempSheet, sheet2)
        self.__StyleSheet(sheet1, headingsStyle, cellStyle)
        self.__StyleSheet(sheet2, headingsStyle, cellStyle)
        book.remove(tempSheet)

        for cell in sheet2["C"]:
            cell.border = Border(top=Side(style=None),
                                 bottom=Side(style=None))
        for cell in sheet2["E"]:
            cell.number_format = '0.00%'
        book.save("report.xlsx")

    def __CreateVerticalBars(self, ax, title, data1, data2, label1, label2, rotation):
        """
        Формирует вертикальную диаграмму для двух данных

        Args:
            ax (axes.SubplotBase): График
            title (str): Заголовок диаграммы
            data1 (dict): Общие данные
            data2 (dict): Данные для выбранной профессии
            label1 (str): Надпись легенды для общих данных
            label2 (str): Надпись легенды для данные выбранной профессии
            rotation (int): Угол поворота надписей оси X

        Returns:
            axes.SubplotBase: График
        """
        xIndexes = np.arange(len(data1.keys()))
        width = 0.35
        ax.set_title(title)
        ax.bar(xIndexes - width / 2, data1.values(), width, label=label1)
        ax.bar(xIndexes + width / 2, data2.values(), width, label=label2)
        ax.legend()
        ax.grid(axis="y", visible=True)
        ax.set_xticks(xIndexes, data1.keys(), rotation=rotation)
        return ax

    def __CreateHorizontalBar(self, ax, title, data):
        """
        Формирует горизонтальную диаграмму

        Args:
            ax (axes.SubplotBase): График
            title (str): Заголовок диаграммы
            data (dict): Данные

        Returns:
            axes.SubplotBase: График
        """
        width = 0.35
        ax.set_title(title)
        ax.barh(list(data.keys())[:10], list(data.values())[:10], width)
        ax.grid(axis="x", visible=True)
        ax.invert_yaxis()
        return ax

    def __CreatePie(self, ax, title, data):
        """
        Формирует круговую диаграмму
        Args:
            ax (axes.SubplotBase): График
            title (str): Заголовок диаграммы
            data (dict): Данные

        Returns:
            axes.SubplotBase: График
        """
        plt.rc('font', size=6)
        ax.set_title(title)
        ax.pie(data.values(), labels=data.keys(), labeldistance=1.1, startangle=-30)
        return ax

    def GenerateImage(self, listData):
        """
        Формирует изображение "graph.png" со статистикой вакансий в виде графиков

        Args:
            listData: Данные файла
        """
        plt.rc('font', size=8)
        figure = plt.figure()
        ax1 = figure.add_subplot(2, 2, 1)
        ax2 = figure.add_subplot(2, 2, 2)
        ax1 = self.__CreateVerticalBars(ax1, "Уровень зарплат по годам", listData[0], listData[1], "средняя з/п",
                                        f'з/п {self.vacancyName}', 90)
        ax2 = self.__CreateVerticalBars(ax2, "Уровень зарплат по годам", listData[2], listData[3],
                                        "Количество вакансий", f'Количество вакансий {self.vacancyName}', 90)
        ax3 = figure.add_subplot(2, 2, 3)
        ax3 = self.__CreateHorizontalBar(ax3, "Уровень зарплат по городам", listData[4])
        plt.rc('font', size=6)
        tempDict = {k: v for k, v in list(listData[5].items())[:10]}
        tempDict["Другие"] = sum(list(listData[5].values())[10:])
        tempDict = dict(sorted(tempDict.items(), key=lambda x: x[1]))
        ax4 = figure.add_subplot(2, 2, 4)
        ax4 = self.__CreatePie(ax4, "Доля вакансий по городам", tempDict)
        plt.tight_layout()
        plt.savefig("graph.png")
        plt.show()

    def GeneratePDF(self, listData):
        """
        Формирует отчет  "report.pdf" со всей статистикой в виде графиков и таблицы

        Args:
            listData: Данные файла:
        """

        self.GenerateExcel(listData)
        self.GenerateImage(listData)

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdfTemplate.html")
        headingsByYear = ["Год", "Средняя зарплата", f'Средняя зарплата - {self.vacancyName}', "Количество вакансий",
                          f'Количество вакансий - {self.vacancyName}']
        headingsByCity = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        pdfTemplate = template.render({
            "fileName": "graph.png",
            "vacancyName": self.vacancyName,
            "headingsByYear": headingsByYear,
            "headingsByCity": headingsByCity,
            "dynamicsSalaries": listData[0],
            "dynamicsSalariesAtVacancy": listData[1],
            "dynamicsCountVacancies": listData[2],
            "dynamicsCountVacanciesAtVacancy": listData[3],
            "citiesSalaryLevel": {k: v for k, v in list(listData[4].items())[:10]},
            "citiesRatioVacancies": {k: f'{round(v * 100, 2)}%' for k, v in list(listData[5].items())[:10]}
        })

        options = {'enable-local-file-access': None}
        config = pdfkit.configuration(wkhtmltopdf=r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe")
        pdfkit.from_string(pdfTemplate, "report.pdf", configuration=config, options=options)

# inputData = InputConnect()
# dataSet = DataSet(inputData.fileName, inputData.vacancyName)
# inputData.PrintData(dataSet)
#
# reportData = Report(dataSet.vacancyNameParameter)
# reportData.GeneratePDF(inputData.GetListData((dataSet)))
