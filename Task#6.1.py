import sys
import csv
import os
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, NamedStyle, Side, Border


class Vacancy:
    def __init__(self, name, salary, areaName, publishedAt):
        self.name, self.salary, self.areaName, self.publishedAt = name, salary, areaName, publishedAt


class DataSet:
    correctFields = ["name", "salary_from", "area_name", "published_at"]

    def __init__(self, fileName, vacancyNameParameter):
        self.fileName = fileName
        self.vacancyNameParameter = vacancyNameParameter
        self.__UniversalParserCSV(fileName)
        self.vacanciesByYear = self.__VacancyFilterByYear()
        self.vacancyByYear = self.__VacancyFilterByYear(self.vacancyNameParameter)
        self.vacanciesByArea = self.__VacancyFilterByArea()

    def __UniversalParserCSV(self, fileName):
        fileReader, columnNames = self.__CsvReader(fileName)
        self.vacanciesObjects = self.__CsvFilter(fileReader, columnNames)

    def __CsvReader(self, fileName):
        file = open(fileName, encoding='utf-8-sig', newline='')
        if os.stat(fileName).st_size == 0:
            print("Пустой файл")
            sys.exit()
        fileReader = csv.DictReader(file)
        columnNames = fileReader.fieldnames
        return fileReader, columnNames

    def __CsvFilter(self, fileReader, columnNames):
        vacancies = []
        columnsCount = len(columnNames)
        for row in fileReader:
            if all(row.values()) and columnsCount == len(row):
                tempRow = {name: row[name] for name in columnNames}
                tempRow['salary_from'] = Salary(tempRow['salary_from'], tempRow.pop('salary_to'),
                                                tempRow.pop("salary_currency"))
                vacancies.append(Vacancy(*[tempRow[key] for key in self.correctFields]))
        return vacancies

    def DynamicsSalaries(self):
        dinamicsSalaries = self.vacanciesByYear.copy()
        for dataByYear in dinamicsSalaries.items():
            averagesByYear = [vacancy.salary.GetAverage() for vacancy in dataByYear[1]]
            dinamicsSalaries[dataByYear[0]] = int(sum(averagesByYear) / len(averagesByYear))
        return dinamicsSalaries

    def DynamicsCountVacancies(self):
        vacancyCountByYear = self.vacanciesByYear.copy()
        for dataByYear in vacancyCountByYear.items():
            vacancyCountByYear[dataByYear[0]] = len(dataByYear[1])
        return vacancyCountByYear

    def DynamicsSalariesAtVacancy(self):
        dinamicsSalaries = self.vacancyByYear.copy()
        if not dinamicsSalaries:
            return {key: 0 for key in self.currentKeys}
        for dataByYear in dinamicsSalaries.items():
            averagesByYear = [vacancy.salary.GetAverage() for vacancy in dataByYear[1]]
            dinamicsSalaries[dataByYear[0]] = int(sum(averagesByYear) / len(averagesByYear))
        return dinamicsSalaries

    def DynamicsCountVacanciesAtVacancy(self):

        vacancyCountByYear = self.vacancyByYear.copy()
        if not vacancyCountByYear:
            return {key: 0 for key in self.currentKeys}
        for dataByYear in vacancyCountByYear.items():
            vacancyCountByYear[dataByYear[0]] = len(dataByYear[1])
        return vacancyCountByYear

    def CitiesSalaryLevel(self):
        vacanciesByArea = self.vacanciesByArea.copy()
        for dataByArea in vacanciesByArea.items():
            averagesByArea = [vacancy.salary.GetAverage() for vacancy in dataByArea[1]]
            vacanciesByArea[dataByArea[0]] = int(sum(averagesByArea) / len(averagesByArea))
        vacanciesByArea = list(
            {k: v for k, v in sorted(vacanciesByArea.items(), key=lambda item: item[1], reverse=True)}.items())
        vacanciesByArea = {items[0]: items[1] for items in vacanciesByArea[:10]}
        return vacanciesByArea

    def CitiesRatioVacancies(self):
        vacanciesByArea = self.vacanciesByArea.copy()
        for dataByArea in vacanciesByArea.items():
            vacanciesByArea[dataByArea[0]] = round(len(dataByArea[1]) / len(self.vacanciesObjects), 4)
        vacanciesByArea = list(
            {k: v for k, v in sorted(vacanciesByArea.items(), key=lambda item: item[1], reverse=True)}.items())
        vacanciesByArea = {items[0]: items[1] for items in vacanciesByArea[:10]}
        return vacanciesByArea

    def __VacancyFilterByYear(self, vacancyName=None):
        dinamicsSalaries = {}
        self.currentKeys = []
        for vacancy in self.vacanciesObjects:
            year = datetime.strptime(vacancy.publishedAt, '%Y-%m-%dT%H:%M:%S%z').year
            self.currentKeys.append(year)
            if vacancyName is None or vacancyName in vacancy.name:
                dinamicsSalaries.setdefault(year, []).append(vacancy)
        return dinamicsSalaries

    def __VacancyFilterByArea(self):
        vacanciesByArea = {}
        for vacancy in self.vacanciesObjects:
            vacanciesByArea.setdefault(vacancy.areaName, []).append(vacancy)
        vacanciesByArea = self.__ClearByArea(vacanciesByArea)

        return vacanciesByArea

    def __ClearByArea(self, vacanciesByArea):
        tempAreas = vacanciesByArea.copy()
        for keyArea in vacanciesByArea.keys():
            if len(vacanciesByArea[keyArea]) / len(self.vacanciesObjects) < 0.01:
                tempAreas.pop(keyArea)
        return tempAreas


class Salary:
    currencyToRub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, salaryFrom, salaryTo, salaryCurrency):
        self.salaryFrom, self.salaryTo, self.salaryCurrency = salaryFrom, salaryTo, salaryCurrency

    def ChangeCurrency(self, salary):
        return salary * self.currencyToRub[self.salaryCurrency]

    def GetAverage(self):
        return self.ChangeCurrency((int(float(self.salaryFrom)) + int(float(self.salaryTo))) / 2)


class InputConnect:
    __requests = {"Введите название файла: ": lambda fileName: fileName,
                  "Введите название профессии: ": lambda vacancyName: vacancyName}

    _responses = {"Динамика уровня зарплат по годам: ": lambda dataSet: dataSet.DynamicsSalaries(),
                  "Динамика уровня зарплат по годам для выбранной профессии: ": lambda
                      dataSet: dataSet.DynamicsSalariesAtVacancy(),
                  "Динамика количества вакансий по годам: ": lambda dataSet: dataSet.DynamicsCountVacancies(),
                  "Динамика количества вакансий по годам для выбранной профессии: ": lambda
                      dataSet: dataSet.DynamicsCountVacanciesAtVacancy(),
                  "Уровень зарплат по городам (в порядке убывания): ": lambda dataSet: dataSet.CitiesSalaryLevel(),
                  "Доля вакансий по городам (в порядке убывания): ": lambda dataSet: dataSet.CitiesRatioVacancies()}

    def __init__(self):
        self.fileName, self.vacancyName = self.__GetData()

    def __GetData(self):
        data = {}
        for request in self.__requests.keys():
            data[request] = input(request)
        return data.values()

    def PrintData(self, dataSet):
        for response in self._responses.items():
            print(f'{response[0]}{response[1](dataSet)}')

    def GetListData(self, dataSet):
        data = []
        for response in self._responses.items():
            data.append(response[1](dataSet))
        return data


class Report:
    def __init__(self, vacancyName):
        self.vacancyName = vacancyName

    def CheckEmptyText(self, value):
        if value is None:
            return ""
        return str(value)

    def __CompleteSheet(self, sheet, data, indexesData, count=None):
        i, j = indexesData
        if count is None:
            count = len(data[i].keys()) + 1
        for year in list(data[i].keys())[:count]:
            row = [year]
            for dictData in data[i:j]:
                row.append(dictData[year])
            sheet.append(row)
        return sheet

    def __CopySheetToSheet(self, sheet1, sheet2):
        maxColumn = sheet2.max_column
        for i in range(1, sheet1.max_row + 1):
            for j in range(1, sheet1.max_column + 1):
                sheet2.cell(row=i + 1, column=maxColumn + j - 2).value = sheet1.cell(row=i, column=j).value
        return sheet2

    def __StyleSheet(self, sheet, headings, headingsStyle, cellStyle):

        for row in sheet.rows:
            for cell in row:
                cell.style = cellStyle
        for cell in sheet["1:1"]:
            cell.style = headingsStyle
        for column in sheet.columns:
            length = max(len(self.CheckEmptyText(cell.value)) for cell in column)
            sheet.column_dimensions[column[0].column_letter].width = length + 2

    def GenerateExcel(self, listData):
        book = openpyxl.Workbook()
        book.remove(book.active)
        sheet1 = book.create_sheet("Статистика по годам")
        sheet2 = book.create_sheet("Статистика по городам")
        headingsByYear = ["Год", "Средняя зарплата", f'Средняя зарплата - {self.vacancyName}', "Количество вакансий",
                          f'Количество вакансий - {self.vacancyName}']
        headingsByCity = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        sheet1.append(headingsByYear)
        sheet2.append(headingsByCity)

        headingsStyle = NamedStyle(name='headingsStyle')
        headingsStyle.font = Font(bold=True)
        headingsStyle.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                      bottom=Side(style='thin'))
        cellStyle = NamedStyle(name='cellStyle')
        cellStyle.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                  bottom=Side(style='thin'))

        tempSheet = book.create_sheet("tempSheet")
        sheet1 = self.__CompleteSheet(sheet1, listData, (0, 4))
        sheet2 = self.__CompleteSheet(sheet2, listData, (4, 5), 10)
        tempSheet = self.__CompleteSheet(tempSheet, listData, (5, 6), 10)
        sheet2 = self.__CopySheetToSheet(tempSheet, sheet2)
        book.remove(tempSheet)

        self.__StyleSheet(sheet1, headingsByYear, headingsStyle, cellStyle)
        self.__StyleSheet(sheet2, headingsByCity, headingsStyle, cellStyle)

        for cell in sheet2["C"]:
            cell.border = Border(top=Side(style=None),
                                 bottom=Side(style=None))
        for cell in sheet2["E"]:
            cell.number_format = '0.00%'
        book.save("report.xlsx")


inputData = InputConnect()
dataSet = DataSet(inputData.fileName, inputData.vacancyName)
inputData.PrintData(dataSet)

reportData = Report(dataSet.vacancyNameParameter)
reportData.GenerateExcel(inputData.GetListData(dataSet))
